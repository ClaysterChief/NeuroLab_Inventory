from io import BytesIO
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Bitacora, Cajas, Ratas

# ── Colores del sistema ────────────────────────────────────────────────────────
DARK   = colors.HexColor('#1a1a2e')
RED    = colors.HexColor('#80201d')
LIGHT  = colors.HexColor('#f8f8fb')
BORDER = colors.HexColor('#e0e0e0')

# ── Estilos de párrafo ─────────────────────────────────────────────────────────
STYLE_TITLE = ParagraphStyle(
    'NLTitle', fontName='Helvetica-Bold', fontSize=16,
    textColor=DARK, spaceAfter=4,
)
STYLE_SUB = ParagraphStyle(
    'NLSub', fontName='Helvetica', fontSize=9,
    textColor=colors.grey, spaceAfter=14,
)
STYLE_SECTION = ParagraphStyle(
    'NLSection', fontName='Helvetica-Bold', fontSize=11,
    textColor=RED, spaceBefore=10, spaceAfter=6,
)


def _table_style():
    return TableStyle([
        # Cabecera
        ('BACKGROUND',    (0, 0), (-1, 0), DARK),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 8),
        ('LINEBELOW',     (0, 0), (-1, 0), 1.5, RED),
        # Cuerpo
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT]),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        # General
        ('GRID',          (0, 0), (-1, -1), 0.4, BORDER),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
    ])


def _fmt(d):
    """Formatea fechas; devuelve '—' si es None."""
    if not d:
        return '—'
    return d.strftime('%d/%m/%Y')


def _trunc(text, n=45):
    """Trunca texto largo para que no rompa la tabla."""
    if not text:
        return '—'
    return text[:n] + '…' if len(text) > n else text


# ── Reporte de Inventario ──────────────────────────────────────────────────────

def generate_inventario_pdf():
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        leftMargin=0.6*inch, rightMargin=0.6*inch,
        topMargin=0.6*inch, bottomMargin=0.6*inch,
        title='Reporte de Inventario — NeuroLab',
    )
    now = datetime.now().strftime('%d/%m/%Y  %H:%M')
    items = []

    items.append(Paragraph('NeuroLab Inventory', STYLE_TITLE))
    items.append(Paragraph(f'Reporte de Inventario  ·  Generado: {now}', STYLE_SUB))

    # ── Cajas ──────────────────────────────────────────────────────────────────
    items.append(Paragraph('Inventario de Cajas', STYLE_SECTION))

    cajas = Cajas.objects.select_related('idusuario').order_by('idcaja')
    rows = [['Caja #', 'Cantidad', 'Sexo', 'F. Nacimiento', 'Talla', 'Responsable', 'Comentarios']]
    for c in cajas:
        rows.append([
            str(c.idcaja),
            str(c.cantidadratas),
            c.sexo or '—',
            _fmt(c.fechanacimiento),
            c.talla or '—',
            c.idusuario.nombreusuario if c.idusuario else '—',
            _trunc(c.comentarios),
        ])
    if len(rows) == 1:
        rows.append(['Sin registros'] + [''] * 6)

    tbl = Table(
        rows,
        colWidths=[0.55*inch, 0.65*inch, 0.75*inch, 0.9*inch,
                   0.65*inch, 1.1*inch, 2.6*inch],
    )
    tbl.setStyle(_table_style())
    items.append(tbl)
    items.append(Spacer(1, 0.2*inch))

    # ── Ratas ──────────────────────────────────────────────────────────────────
    items.append(Paragraph('Registro de Ratas', STYLE_SECTION))

    ratas = Ratas.objects.select_related(
        'idcondicion', 'idcaja'
    ).order_by('sexo', 'idrata')
    rows2 = [['ID', 'Sexo', 'N° Cola', 'Caja', 'Condición', 'Peso semanal (g)', 'F. Cirugía']]
    for r in ratas:
        prefix = r.sexo[0] if r.sexo else '?'
        rows2.append([
            f'{prefix}-{r.idrata}',
            r.sexo or '—',
            str(r.numerocola),
            f'Caja #{r.idcaja.idcaja}' if r.idcaja else '—',
            r.idcondicion.nombrecondicion if r.idcondicion else '—',
            str(r.pesosemanal) if r.pesosemanal is not None else '—',
            _fmt(r.fechacirugia),
        ])
    if len(rows2) == 1:
        rows2.append(['Sin registros'] + [''] * 6)

    tbl2 = Table(
        rows2,
        colWidths=[0.55*inch, 0.75*inch, 0.65*inch, 0.8*inch,
                   1.2*inch, 1.15*inch, 0.9*inch],
    )
    tbl2.setStyle(_table_style())
    items.append(tbl2)

    doc.build(items)
    buffer.seek(0)
    return buffer


# ── Reporte de Bitácora ────────────────────────────────────────────────────────

def generate_bitacora_pdf():
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        leftMargin=0.6*inch, rightMargin=0.6*inch,
        topMargin=0.6*inch, bottomMargin=0.6*inch,
        title='Reporte de Bitácora — NeuroLab',
    )
    now = datetime.now().strftime('%d/%m/%Y  %H:%M')
    items = []

    items.append(Paragraph('NeuroLab Inventory', STYLE_TITLE))
    items.append(Paragraph(f'Reporte de Bitácora Experimental  ·  Generado: {now}', STYLE_SUB))
    items.append(Paragraph('Bitácora de Experimentos', STYLE_SECTION))

    registros = Bitacora.objects.select_related(
        'idrata', 'idusuario', 'idanestesico', 'idtejido'
    ).order_by('idbitacora')

    rows = [['#', 'Rata', 'Fecha', 'Anestésico', 'Dosis total (ml)',
             'Peso exp. (g)', 'Tejido', 'Responsable', 'Actividad']]
    for b in registros:
        rata = b.idrata
        prefix = rata.sexo[0] if rata and rata.sexo else '?'
        rata_label = f'{prefix}-{rata.idrata}' if rata else '—'
        rows.append([
            str(b.idbitacora),
            rata_label,
            _fmt(b.fechacirujia),
            b.idanestesico.nombreanestesico if b.idanestesico else '—',
            str(b.dosistotal)     if b.dosistotal     is not None else '—',
            str(b.pesoexperimento) if b.pesoexperimento is not None else '—',
            b.idtejido.nombretejido if b.idtejido else '—',
            b.idusuario.nombreusuario if b.idusuario else '—',
            _trunc(b.actividad, 50),
        ])
    if len(rows) == 1:
        rows.append(['Sin registros'] + [''] * 8)

    tbl = Table(
        rows,
        colWidths=[0.35*inch, 0.55*inch, 0.8*inch, 1.1*inch, 0.9*inch,
                   0.9*inch, 0.9*inch, 1.0*inch, 2.7*inch],
    )
    tbl.setStyle(_table_style())
    items.append(tbl)

    doc.build(items)
    buffer.seek(0)
    return buffer

# ── Colores Excel ──────────────────────────────────────────────────────────────
XL_DARK  = '1a1a2e'
XL_RED   = '80201d'
XL_LIGHT = 'f8f8fb'


def _xl_header_style(cell):
    cell.font      = Font(bold=True, color='FFFFFF', size=10)
    cell.fill      = PatternFill('solid', fgColor=XL_DARK)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = Border(
        bottom=Side(style='medium', color=XL_RED)
    )


def _xl_cell_style(cell, even=True):
    cell.fill      = PatternFill('solid', fgColor=XL_LIGHT if even else 'FFFFFF')
    cell.font      = Font(size=9)
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    cell.border    = Border(
        bottom=Side(style='thin', color='e0e0e0'),
        right=Side(style='thin',  color='e0e0e0'),
    )


def _xl_title(ws, title, subtitle):
    ws.merge_cells('A1:H1')
    t = ws['A1']
    t.value     = title
    t.font      = Font(bold=True, size=14, color=XL_DARK)
    t.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:H2')
    s = ws['A2']
    s.value     = subtitle
    s.font      = Font(size=9, color='888888')
    s.alignment = Alignment(horizontal='left')
    ws.row_dimensions[2].height = 16


def generate_inventario_excel():
    wb = Workbook()
    now_str = datetime.now().strftime('%d/%m/%Y %H:%M')

    # ── Hoja 1: Cajas ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Cajas'
    _xl_title(ws1, 'NeuroLab Inventory — Inventario de Cajas', f'Generado: {now_str}')

    headers1 = ['Caja #', 'Cantidad ratas', 'Sexo', 'F. Nacimiento', 'Talla', 'Responsable', 'Comentarios']
    widths1  = [9, 14, 10, 14, 10, 18, 30]
    row_start = 4

    for col, (h, w) in enumerate(zip(headers1, widths1), 1):
        cell = ws1.cell(row=row_start, column=col, value=h)
        _xl_header_style(cell)
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[row_start].height = 20

    cajas = Cajas.objects.select_related('idusuario').order_by('idcaja')
    for i, c in enumerate(cajas, 1):
        row = [
            c.idcaja, c.cantidadratas, c.sexo or '—',
            _fmt(c.fechanacimiento), c.talla or '—',
            c.idusuario.nombreusuario if c.idusuario else '—',
            c.comentarios or '—',
        ]
        for col, val in enumerate(row, 1):
            cell = ws1.cell(row=row_start + i, column=col, value=val)
            _xl_cell_style(cell, i % 2 == 0)

    # ── Hoja 2: Ratas ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Ratas')
    _xl_title(ws2, 'NeuroLab Inventory — Registro de Ratas', f'Generado: {now_str}')

    headers2 = ['ID', 'Sexo', 'N° Cola', 'Caja', 'Condición', 'Peso semanal (g)', 'F. Cirugía']
    widths2  = [8, 10, 10, 12, 16, 17, 13]

    for col, (h, w) in enumerate(zip(headers2, widths2), 1):
        cell = ws2.cell(row=row_start, column=col, value=h)
        _xl_header_style(cell)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[row_start].height = 20

    ratas = Ratas.objects.select_related('idcondicion', 'idcaja').order_by('sexo', 'idrata')
    for i, r in enumerate(ratas, 1):
        prefix = r.sexo[0] if r.sexo else '?'
        row = [
            f'{prefix}-{r.idrata}', r.sexo or '—', r.numerocola,
            f'Caja #{r.idcaja.idcaja}' if r.idcaja else '—',
            r.idcondicion.nombrecondicion if r.idcondicion else '—',
            r.pesosemanal if r.pesosemanal is not None else '—',
            _fmt(r.fechacirugia),
        ]
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=row_start + i, column=col, value=val)
            _xl_cell_style(cell, i % 2 == 0)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_bitacora_excel():
    wb = Workbook()
    now_str = datetime.now().strftime('%d/%m/%Y %H:%M')

    ws = wb.active
    ws.title = 'Bitácora'
    _xl_title(ws, 'NeuroLab Inventory — Bitácora de Experimentos', f'Generado: {now_str}')

    headers = ['#', 'Rata', 'Fecha', 'Anestésico', 'Dosis total (ml)',
               'Peso exp. (g)', 'Tejido', 'Responsable', 'Actividad', 'Notas']
    widths  = [6, 8, 12, 16, 15, 13, 16, 16, 35, 25]
    row_start = 4

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row_start, column=col, value=h)
        _xl_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row_start].height = 20

    registros = Bitacora.objects.select_related(
        'idrata', 'idusuario', 'idanestesico', 'idtejido'
    ).order_by('idbitacora')

    for i, b in enumerate(registros, 1):
        rata   = b.idrata
        prefix = rata.sexo[0] if rata and rata.sexo else '?'
        row = [
            b.idbitacora,
            f'{prefix}-{rata.idrata}' if rata else '—',
            _fmt(b.fechacirujia),
            b.idanestesico.nombreanestesico if b.idanestesico else '—',
            b.dosistotal      if b.dosistotal      is not None else '—',
            b.pesoexperimento if b.pesoexperimento is not None else '—',
            b.idtejido.nombretejido if b.idtejido else '—',
            b.idusuario.nombreusuario if b.idusuario else '—',
            b.actividad or '—',
            b.notas     or '—',
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_start + i, column=col, value=val)
            _xl_cell_style(cell, i % 2 == 0)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer